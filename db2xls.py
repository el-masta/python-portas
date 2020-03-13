#!/usr/bin/env python
# -*- coding: utf-8 -*-
#
#  portas.py
#  
#  Mario Rodríguez
#  
#  Crea una hoja de excel con los datos de portabilidades del día.
#  
#  Requiere libreria openpyxl


def main(args):
   return 0

if __name__ == '__main__':
   import sys
   import mysql.connector as mariadb
   from datetime import datetime, date, time, timedelta
   import calendar
   from openpyxl import Workbook
   from openpyxl.styles import NamedStyle, Font, Border, Side, PatternFill, Alignment
   
   
   #Define conexión
   conn = mariadb.connect(host='localhost', port='3306', user='*****', password='*****', database='cisat')
   cursor = conn.cursor()
   fecha=date.today()
   #fecha=date(2019,10,18)
   dia=fecha.strftime("%Y-%m-%d")
   #Conecta y executa
   try:
      #Añade TODAS las portas
      cursor.execute("SELECT telefono, concat(curps.nombre,' ',curps.apep,' ',curps.apem) as nombre, idcop, iccid, operador, promotor from portas LEFT JOIN curps on curps.curp = portas.curp WHERE agencia='DREAM1' and fecha='"+dia+"' ORDER BY id")
      
      #Añade SOLAMENTE las portas exitosas
      #cursor.execute("SELECT telefono, concat(curps.nombre,' ',curps.apep,' ',curps.apem) as nombre, idcop, iccid, operador, promotor from portas LEFT JOIN curps on curps.curp = portas.curp WHERE agencia='DREAM1' and fecha='"+dia+"' and status='ACEPTADO' ORDER BY id")
      
      records=cursor.fetchall()
      #Crea hoja de calculo
      print(dia)
      xls= Workbook()
      hoja=xls.active
      hoja.title="Dream"

      # Registra estilos
      normal = NamedStyle(name="normal")
      normal.font  = Font(name="Calibri", size=11)
      bd = Side(style='thin', color="000000")
      normal.border = Border(left=bd, top=bd, right=bd, bottom=bd)
      xls.add_named_style(normal)
      
      centrado = NamedStyle(name="centrado")
      centrado.font  = Font(name="Calibri", size=11)
      bd = Side(style='thin', color="000000")
      centrado.border = Border(left=bd, top=bd, right=bd, bottom=bd)
      centrado.alignment = Alignment(horizontal='center')
      xls.add_named_style(centrado)
      
      titulo = NamedStyle(name="titulo")
      titulo.font = Font(name="Calibri",size=26)
      bd = Side(style='thick', color="000000")
      titulo.border = Border(left=bd, top=bd, right=bd, bottom=bd)
      titulo.fill = PatternFill("solid", fgColor="1f3964")
      titulo.alignment = Alignment(horizontal='center',vertical='center')
      xls.add_named_style(titulo)
      
      cabeza = NamedStyle(name="cabeza")
      cabeza.font = Font(name="Calibri", size=11, bold=True)
      bd = Side(style='thin', color="000000")
      cabeza.border = Border(left=bd, top=bd, right=bd, bottom=bd)
      cabeza.fill = PatternFill("solid", fgColor="deebf7")
      cabeza.alignment = Alignment(horizontal='center',vertical='center')
      xls.add_named_style(cabeza)
      
      info = NamedStyle(name="info")
      info.font = Font(name="Calibri", size=11)
      bd = Side(style='thick', color="000000")
      info.border = Border(left=bd, top=bd, right=bd, bottom=bd)
      info.fill = PatternFill("solid", fgColor="fbe4d5")
      info.alignment = Alignment(horizontal='center',vertical='center')
      xls.add_named_style(info)
      
      #Crea encabezados
      hoja.merge_cells('B2:I4')
      hoja['B2']='REPORTE DE PRODUCCIÓN'
      hoja['B2'].style = 'titulo'
      hoja['B6']='SEMANA DE PRODUCCIÓN'
      hoja['B6'].style = 'info'
      hoja['C6']=fecha.isocalendar()[1]
      hoja['C6'].style = 'info'
      hoja['B7']='DÍA REPORTADO'
      hoja['B7'].style = 'info'
      hoja['C7']=fecha.strftime("%d/%m/%Y")
      hoja['C7'].style = 'info'
      hoja['B8']='TOTAL'
      hoja['B8'].style = 'info'
      hoja['B8'].fill = PatternFill("solid", fgColor="e2efd9")
      hoja['C8']=len(records)
      hoja['C8'].style = 'info'
      hoja['C8'].fill = PatternFill("solid", fgColor="e2efd9")
      hoja['C8'].font = Font(name="Calibri", size=11, bold=True)
      
      hoja['B10']='NÚMERO PORTADO'
      hoja['B10'].style = 'cabeza'
      hoja['C10']='NOMBRE DEL CLIENTE'
      hoja['C10'].style= 'cabeza'
      hoja['D10']='FECHA PORT IN'
      hoja['D10'].style= 'cabeza'
      hoja['E10']='ID COP'
      hoja['E10'].style= 'cabeza'
      hoja['F10']='# SIM ASIGNADO'
      hoja['F10'].style= 'cabeza'
      hoja['G10']='DONADOR'
      hoja['G10'].style= 'cabeza'
      hoja['H10']='NOMBRE DE PROMOTOR'
      hoja['H10'].style= 'cabeza'
      hoja['I10']='NOMBRE DE SUPERVISOR'
      hoja['I10'].style= 'cabeza'


      row=11
      datos=[]
      #Llena celdas con datos de la porta
      
      for TELEFONO, NOMBRE, IDCOP, ICCID, OPERADOR, PROMOTOR in records:
         datos.append((TELEFONO, NOMBRE, IDCOP, ICCID, OPERADOR, PROMOTOR))
         
      #Ordena los datos por USUARIO de app
      #datos.sort(key=lambda tup: tup[6][12::])
      datos.sort(key=lambda tup: tup[5])
         
      for TELEFONO, NOMBRE, IDCOP, ICCID, OPERADOR, PROMOTOR in datos:
         hoja['B'+str(row)].style= 'normal'
         hoja['C'+str(row)].style= 'normal'
         hoja['D'+str(row)].style= 'normal'
         hoja['E'+str(row)].style= 'normal'
         hoja['F'+str(row)].style= 'normal'
         hoja['G'+str(row)].style= 'centrado'
         hoja['H'+str(row)].style= 'centrado'
         hoja['I'+str(row)].style= 'centrado'

       
         hoja['B'+str(row)] = TELEFONO
         hoja['C'+str(row)] = NOMBRE
         hoja['D'+str(row)] = fecha.strftime("%d/%m/%Y") #FECHA
         hoja['E'+str(row)] = IDCOP
         hoja['F'+str(row)] = ICCID
         hoja['G'+str(row)] = OPERADOR
         hoja['H'+str(row)] = PROMOTOR
         hoja['I'+str(row)] = 'ISMAEL GOMEZ'

         
         row+=1
      
      #Ajuste de columnas
      for col in hoja.columns:
         max_length = 0
         column = col[0].column # Obtiene nombre de columna
         for cell in col:
            if cell.coordinate in hoja.merged_cells: # Checa celdas combinadas
               continue
            try: # Evita error en celdas vacias
               if len(str(cell.value)) > max_length:
                  max_length = len(cell.value)
            except:
               pass
         adjusted_width = (max_length + 6) 
         hoja.column_dimensions[column].width = adjusted_width
      hoja.column_dimensions['A'].width = 2
      hoja.column_dimensions['B'].width = 25

      #Guarda archivo
      xls.save("/home/mario/Escritorio/DREAM/"+fecha.strftime("%d-%m-%Y")+"-Produccion-ISMAEL-GOMEZ.xlsx")
      
   except mariadb.Error as error:
      print("Error2: {}".format(error))
   finally:
      if(conn.is_connected() ):
         conn.close()
      print('Conexión finalizada')

   sys.exit(main(sys.argv))
